import os
from pathlib import Path
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import shutil
import zipfile
import glob
from PyPDF2 import PdfReader
import re
import warnings
from datetime import datetime

# Ignorar advertencias de PyPDF2
warnings.filterwarnings('ignore', category=UserWarning, module='PyPDF2')

# Cargar variables de entorno
load_dotenv()

# Configuración desde variables de entorno
BASE_DIR = Path(__file__).parent.resolve()
PDF_DIR_NAME = os.getenv('PDF_DIR', 'downloads')  # Usar el valor de PDF_DIR del .env
DOWNLOADS_DIR = BASE_DIR / PDF_DIR_NAME
EXCEL_FILE = os.getenv('EXCEL_FILE', 'SEGUIMIENTO_ALMACEN.xlsx')
EXCEL_SHEET = os.getenv('EXCEL_SHEET', 'Hoja1')  # Nombre de la hoja de Excel
MINDRAY_USERNAME = os.getenv('MINDRAY_USERNAME')
MINDRAY_PASSWORD = os.getenv('MINDRAY_PASSWORD')
MINDRAY_URL = os.getenv('MINDRAY_URL', 'https://service.mindray.com/SSO/Login/Signin?app=coa')

# Configuración de columnas Excel
EXCEL_COL_FECHA = os.getenv('EXCEL_COL_FECHA', 'B')  # Fecha
EXCEL_COL_CODIGO = os.getenv('EXCEL_COL_CODIGO', 'F')  # Código
EXCEL_COL_LOTE = os.getenv('EXCEL_COL_LOTE', 'H')  # Lote
EXCEL_COL_FECHA_VENC = os.getenv('EXCEL_COL_FECHA_VENC', 'L')  # Fecha vencimiento
EXCEL_COL_COA = os.getenv('EXCEL_COL_COA', 'M')  # COA
EXCEL_COL_ALMACEN = os.getenv('EXCEL_COL_ALMACEN', 'AH')  # Almacén

def excel_column_to_number(column_letter):
    """
    Convierte una letra de columna de Excel a su número correspondiente (base 1).
    Ejemplos:
        A -> 1
        B -> 2
        Z -> 26
        AA -> 27
        AH -> 34
    """
    result = 0
    for i, letter in enumerate(reversed(column_letter.upper())):
        result += (ord(letter) - ord('A') + 1) * (26 ** i)
    return result

def setup_driver():
    """Configurar el driver de Chrome con las opciones necesarias"""
    # Asegurar que el directorio existe
    DOWNLOADS_DIR.mkdir(exist_ok=True)
    print(f"\nDirectorio de descargas configurado en: {DOWNLOADS_DIR}")
    print(f"Usando carpeta: {PDF_DIR_NAME}")
    
    chrome_options = Options()
    # chrome_options.add_argument('--headless')  # Descomenta para ejecutar sin interfaz
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    
    # Configurar preferencias de descarga
    prefs = {
        "download.default_directory": str(DOWNLOADS_DIR),  # Directorio de descarga
        "download.prompt_for_download": False,  # No mostrar el diálogo de guardar como
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1  # Permitir descargas múltiples
    }
    chrome_options.add_experimental_option("prefs", prefs)
    
    # Configurar el servicio de Chrome
    service = Service(ChromeDriverManager().install())
    
    return webdriver.Chrome(service=service, options=chrome_options)

def change_language_to_english(driver):
    """Cambiar el idioma de la página a inglés usando el menú desplegable"""
    try:
        print("\nCambiando idioma a inglés...")
        time.sleep(3)  # Esperar a que la página esté completamente cargada
        
        # Encontrar el nav y el dropdown de idioma
        wait = WebDriverWait(driver, 10)
        
        # Encontrar y hacer clic en el dropdown de idioma
        lang_dropdown = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//li[contains(@class, 'dropdown-language')]//a[@id='langSelected']"))
        )
        driver.execute_script("arguments[0].click();", lang_dropdown)
        time.sleep(1)
        
        # Encontrar y hacer clic en la opción de inglés
        english_option = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//ul[@id='ulChooseLang']//li/a[@value='en-us']"))
        )
        driver.execute_script("arguments[0].click();", english_option)
        
        print("Idioma cambiado a inglés")
        time.sleep(2)
        
    except Exception as e:
        print(f"Error al cambiar el idioma: {str(e)}")
        print("Detalles del error:")
        try:
            print(f"URL actual: {driver.current_url}")
            nav = driver.find_element(By.CLASS_NAME, "navbar-nav")
            print("Elementos en el nav:")
            print(nav.get_attribute('innerHTML'))
        except Exception as detail_error:
            print(f"Error al obtener detalles: {str(detail_error)}")

def login_mindray(driver):
    """Iniciar sesión en el portal de Mindray"""
    try:
        print("\n=== INICIANDO SESIÓN EN MINDRAY ===")
        
        # Cargar la página de inicio de sesión
        print(f"Accediendo a {MINDRAY_URL}")
        driver.get(MINDRAY_URL)
        
        # Esperar a que la página cargue completamente
        time.sleep(3)
        
        # Configurar espera explícita
        wait = WebDriverWait(driver, 20)
        
        # Encontrar e ingresar el nombre de usuario usando el ID exacto
        print("Buscando campo de usuario...")
        username_field = wait.until(
            EC.presence_of_element_located((By.ID, "txtUsername"))
        )
        username_field.clear()
        username_field.send_keys(MINDRAY_USERNAME)
        print("Usuario ingresado")
        
        # Encontrar e ingresar la contraseña usando el ID exacto
        print("Buscando campo de contraseña...")
        password_field = wait.until(
            EC.presence_of_element_located((By.ID, "txtPassword"))
        )
        password_field.clear()
        password_field.send_keys(MINDRAY_PASSWORD)
        print("Contraseña ingresada")
        
        # Encontrar y hacer clic en el botón de inicio de sesión usando el ID exacto
        print("Buscando botón de inicio de sesión...")
        login_button = wait.until(
            EC.element_to_be_clickable((By.ID, "btnLogin"))
        )
        
        # Esperar un momento antes de hacer clic
        time.sleep(2)
        print("Haciendo clic en el botón de inicio de sesión...")
        login_button.click()
        
        # Esperar a que se complete el inicio de sesión
        time.sleep(5)
        
        # Verificar si el inicio de sesión fue exitoso
        if "Login" not in driver.title:
            print("Inicio de sesión exitoso")
            # Cambiar el idioma a inglés después del login exitoso
            change_language_to_english(driver)
            return True
        else:
            print("Error: No se pudo iniciar sesión")
            return False
            
    except Exception as e:
        print(f"\nError durante el inicio de sesión: {str(e)}")
        print("\nDetalles del estado actual:")
        try:
            print(f"URL actual: {driver.current_url}")
            print(f"Título de la página: {driver.title}")
            print("\nEstado del formulario:")
            form = driver.find_element(By.CLASS_NAME, "login-form")
            print("Formulario encontrado")
            print("Campos visibles:")
            print(f"- Usuario: {driver.find_element(By.ID, 'txtUsername').is_displayed()}")
            print(f"- Contraseña: {driver.find_element(By.ID, 'txtPassword').is_displayed()}")
            print(f"- Botón login: {driver.find_element(By.ID, 'btnLogin').is_displayed()}")
        except Exception as form_error:
            print(f"Error al inspeccionar el formulario: {str(form_error)}")
        return False

def clean_material_code(code):
    """Limpia el código de material eliminando 'MR ' si existe"""
    if pd.isna(code):
        return ''
    code_str = str(code)
    return code_str.replace('MR ', '')

def read_excel_data():
    """Leer datos del Excel necesarios para la búsqueda"""
    try:
        print("\n=== LEYENDO ARCHIVO EXCEL ===")
        
        # Verificar archivo
        excel_path = BASE_DIR / EXCEL_FILE
        if not excel_path.exists():
            raise FileNotFoundError(f"No se encontró el archivo Excel en: {excel_path}")
            
        print(f"Archivo Excel encontrado: {EXCEL_FILE}")
        print(f"Leyendo hoja: {EXCEL_SHEET}")
        
        # Convertir letras de columnas a números (restamos 1 porque pandas usa base 0)
        cols = [
            excel_column_to_number(EXCEL_COL_FECHA) - 1,    # Fecha
            excel_column_to_number(EXCEL_COL_CODIGO) - 1,   # Código
            excel_column_to_number(EXCEL_COL_LOTE) - 1,     # Lote
            excel_column_to_number(EXCEL_COL_COA) - 1,      # COA
            excel_column_to_number(EXCEL_COL_ALMACEN) - 1,  # Almacén
        ]
        
        # Leer Excel seleccionando solo las columnas necesarias
        df = pd.read_excel(
            excel_path,
            sheet_name=EXCEL_SHEET,  # Especificar la hoja a leer
            usecols=cols,
            engine='openpyxl'
        )
        
        # Renombrar las columnas para trabajar más fácilmente
        df.columns = ['FECHA', 'CODIGO', 'LOTE', 'COA', 'ALMACEN']
        
        # Filtrar registros para búsqueda (año 2025, sin COA y no PRODIS)
        df_search = df[
            (pd.to_datetime(df['FECHA']).dt.year == 2025) & 
            (df['COA'].isna() | (df['COA'].astype(str).str.strip() == '')) &
            (df['ALMACEN'] != 'PRODIS')
        ].copy()
        
        # Crear la columna combinada para búsqueda
        df_search['CODIGO_LOTE'] = df_search.apply(
            lambda row: f"{str(row['CODIGO']).replace('MR ', '')} {str(row['LOTE'])}".strip(),
            axis=1
        )
        
        print(f"Total de registros a buscar: {len(df_search)}")
        print("Filtros aplicados:")
        print("- Año: 2025")
        print("- Sin COA")
        print("- Almacén diferente a PRODIS")
        
        return df_search['CODIGO_LOTE'].tolist()
        
    except Exception as e:
        print(f"Error al leer el Excel: {str(e)}")
        raise

def wait_for_results(driver, num_records):
    """Esperar según la cantidad de registros"""
    wait_time = max(15, (num_records // 100) * 15)  # 15 segundos por cada 100 registros
    print(f"Esperando {wait_time} segundos para que carguen {num_records} registros...")
    time.sleep(wait_time)

def process_zip_files():
    """Procesar archivos ZIP: descomprimir y eliminar"""
    try:
        print("\nProcesando archivos ZIP...")
        zip_files = list(DOWNLOADS_DIR.glob("*.zip"))
        
        if not zip_files:
            print("No se encontraron archivos ZIP para procesar")
            return
        
        print(f"Encontrados {len(zip_files)} archivos ZIP")
        
        for zip_file in zip_files:
            try:
                print(f"\nProcesando: {zip_file.name}")
                
                # Descomprimir el archivo ZIP
                with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                    # Listar contenido antes de extraer
                    file_list = zip_ref.namelist()
                    print(f"Contenido del ZIP: {len(file_list)} archivos")
                    
                    # Extraer todos los archivos
                    zip_ref.extractall(DOWNLOADS_DIR)
                    print("Archivos extraídos correctamente")
                
                # Eliminar el archivo ZIP
                zip_file.unlink()
                print(f"Archivo ZIP eliminado: {zip_file.name}")
                
            except Exception as e:
                print(f"Error procesando {zip_file.name}: {str(e)}")
                continue
        
        print("\nProcesamiento de archivos ZIP completado")
        
    except Exception as e:
        print(f"Error durante el procesamiento de ZIPs: {str(e)}")

def clean_downloads_directory():
    """Limpiar la carpeta de descargas antes de comenzar"""
    try:
        print("\nLimpiando directorio de descargas...")
        if DOWNLOADS_DIR.exists():
            # Eliminar todos los archivos en el directorio
            for file in DOWNLOADS_DIR.glob("*"):
                try:
                    if file.is_file():
                        file.unlink()
                        print(f"Archivo eliminado: {file.name}")
                    elif file.is_dir():
                        shutil.rmtree(file)
                        print(f"Carpeta eliminada: {file.name}")
                except Exception as e:
                    print(f"Error al eliminar {file}: {str(e)}")
            print("Limpieza del directorio completada")
        else:
            print("El directorio no existe, se creará uno nuevo")
            DOWNLOADS_DIR.mkdir(exist_ok=True)
    except Exception as e:
        print(f"Error durante la limpieza del directorio: {str(e)}")

def wait_for_download_complete(timeout=300):
    """
    Espera a que se complete la descarga actual.
    Args:
        timeout: Tiempo máximo de espera en segundos (default 5 minutos)
    Returns:
        bool: True si la descarga se completó, False si se agotó el tiempo
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        # Buscar archivos .zip y .part en el directorio de descargas
        zip_files = list(DOWNLOADS_DIR.glob("*.zip"))
        part_files = list(DOWNLOADS_DIR.glob("*.part"))
        crdownload_files = list(DOWNLOADS_DIR.glob("*.crdownload"))
        
        # Si hay archivos .zip y no hay archivos parciales, la descarga está completa
        if zip_files and not part_files and not crdownload_files:
            print("Descarga completada")
            return True
        
        # Si aún hay archivos parciales, esperar
        if part_files or crdownload_files:
            print("Descarga en progreso...")
            time.sleep(5)
            continue
            
        # Si no hay archivos .zip ni archivos parciales, esperar un poco más
        time.sleep(2)
    
    print("Tiempo de espera agotado para la descarga")
    return False

def click_with_retry(driver, element, max_attempts=3, wait_time=2):
    """
    Intenta hacer clic en un elemento con reintentos y diferentes estrategias.
    Args:
        driver: WebDriver de Selenium
        element: Elemento web a hacer clic
        max_attempts: Número máximo de intentos
        wait_time: Tiempo de espera entre intentos
    Returns:
        bool: True si el clic fue exitoso, False si no
    """
    for attempt in range(max_attempts):
        try:
            # Intentar diferentes estrategias de clic
            strategies = [
                # 1. Clic normal con Selenium
                lambda: element.click(),
                # 2. Clic con JavaScript
                lambda: driver.execute_script("arguments[0].click();", element),
                # 3. Clic con JavaScript más forzado
                lambda: driver.execute_script("""
                    var evt = new MouseEvent('click', {
                        bubbles: true,
                        cancelable: true,
                        view: window
                    });
                    arguments[0].dispatchEvent(evt);
                """, element),
                # 4. Scroll al elemento y clic
                lambda: (driver.execute_script("arguments[0].scrollIntoView(true);", element),
                        time.sleep(0.5),
                        element.click())
            ]
            
            # Intentar cada estrategia
            for strategy in strategies:
                try:
                    strategy()
                    time.sleep(wait_time)  # Esperar para ver si funcionó
                    return True
                except:
                    continue
                    
        except Exception as e:
            print(f"Intento {attempt + 1} fallido: {str(e)}")
            time.sleep(wait_time)
    
    return False

def verify_download_button_state(driver, button):
    """
    Verifica el estado del botón de descarga
    """
    try:
        # Verificar si el botón está visible y habilitado
        is_displayed = button.is_displayed()
        is_enabled = button.is_enabled()
        
        # Obtener clases y atributos del botón
        classes = button.get_attribute("class")
        disabled = button.get_attribute("disabled")
        
        print(f"Estado del botón de descarga:")
        print(f"- Visible: {is_displayed}")
        print(f"- Habilitado: {is_enabled}")
        print(f"- Clases: {classes}")
        print(f"- Disabled: {disabled}")
        
        return is_displayed and is_enabled and not disabled
    except Exception as e:
        print(f"Error al verificar estado del botón: {str(e)}")
        return False

def select_and_download_all_files(driver):
    """Seleccionar y descargar todos los archivos"""
    try:
        wait = WebDriverWait(driver, 10)
        
        # Limpiar directorio antes de comenzar las descargas
        clean_downloads_directory()
        
        total_pages = len(driver.find_elements(By.XPATH, "//div[@id='tableReports_paginate']//li[contains(@class, 'paginate_button') and not(contains(@class, 'next')) and not(contains(@class, 'previous')) and not(contains(@class, 'first')) and not(contains(@class, 'last'))]"))
        print(f"Total de páginas encontradas: {total_pages}")
        
        for page in range(1, total_pages + 1):
            print(f"\nProcesando página {page} de {total_pages}")
            
            # Esperar a que la tabla se cargue
            time.sleep(10)
            
            # Primero deseleccionar todo usando JavaScript
            uncheck_script = """
                // Deseleccionar usando iCheck
                var $checkAll = $('#checkAll');
                var $rowCheckboxes = $('.table-btn-check');
                
                // Deseleccionar checkbox principal usando iCheck
                if ($checkAll.length) {
                    $checkAll.iCheck('uncheck');
                }
                
                // Deseleccionar checkboxes de filas usando iCheck
                $rowCheckboxes.iCheck('uncheck');
                
                return true;
            """
            driver.execute_script(uncheck_script)
            time.sleep(3)
            
            # Ahora seleccionar los elementos de la página actual
            check_all_script = """
                // Seleccionar usando iCheck
                var $checkAll = $('#checkAll');
                var $visibleRows = $('#tableReports tbody tr:not(.hidden):not(.filtered)');
                var $visibleCheckboxes = $visibleRows.find('.table-btn-check');
                
                // Seleccionar checkbox principal usando iCheck
                if ($checkAll.length) {
                    $checkAll.iCheck('check');
                    
                    // Seleccionar checkboxes visibles usando iCheck
                    $visibleCheckboxes.iCheck('check');
                    
                    return $visibleRows.length;
                }
                return 0;
            """
            
            selected_count = driver.execute_script(check_all_script)
            print(f"Registros seleccionados en página {page}: {selected_count}")
            time.sleep(3)
            
            # Esperar y verificar el botón de descarga
            print("Esperando botón de descarga...")
            batch_download = None
            try:
                batch_download = wait.until(
                    EC.presence_of_element_located((By.ID, "btnBatchDownload"))
                )
            except Exception as e:
                print(f"Error al encontrar el botón de descarga: {str(e)}")
                continue
            
            # Verificar estado del botón antes de intentar clic
            if not verify_download_button_state(driver, batch_download):
                print("Botón de descarga no está en estado correcto, reintentando...")
                time.sleep(5)
                if not verify_download_button_state(driver, batch_download):
                    print("Error: Botón de descarga no disponible después de reintento")
                    continue
            
            # Intentar hacer clic con reintentos
            print("Intentando hacer clic en el botón de descarga...")
            if not click_with_retry(driver, batch_download):
                print(f"Error: No se pudo hacer clic en el botón de descarga en la página {page}")
                # Intentar refrescar la página y continuar
                driver.refresh()
                time.sleep(10)
                continue
            
            print(f"Descargando archivos de la página {page}")
            
            # Esperar a que se complete la descarga actual
            if not wait_for_download_complete():
                print(f"Error: La descarga de la página {page} no se completó en el tiempo esperado")
            
            # Procesar los archivos ZIP inmediatamente después de la descarga
            process_zip_files()
            
            # Deseleccionar todo después de la descarga
            print("Deseleccionando registros después de la descarga...")
            driver.execute_script(uncheck_script)
            time.sleep(3)
            
            # Si no es la última página, ir a la siguiente
            if page < total_pages:
                print("Navegando a la siguiente página...")
                next_button = wait.until(
                    EC.element_to_be_clickable((By.ID, "tableReports_next"))
                )
                driver.execute_script("arguments[0].click();", next_button)
                time.sleep(8)
        
        print("\nProceso de descarga completado")
        
    except Exception as e:
        print(f"Error durante la selección y descarga: {str(e)}")

def perform_advanced_search(driver, search_codes):
    """Realizar búsqueda avanzada con los códigos y lotes"""
    try:
        print("\nIniciando búsqueda avanzada...")
        wait = WebDriverWait(driver, 10)
        
        # Aceptar descargas múltiples si aparece el diálogo
        try:
            driver.execute_script("""
                window.addEventListener('beforeunload', function(e) {
                    e.preventDefault();
                    return true;
                });
            """)
        except Exception as e:
            print("Nota: No se pudo configurar el manejador de descargas múltiples:", str(e))
        
        # Hacer clic en el botón de búsqueda avanzada
        print("Abriendo ventana de búsqueda avanzada...")
        adv_search_btn = wait.until(
            EC.element_to_be_clickable((By.ID, "btnAdvSearch"))
        )
        driver.execute_script("arguments[0].click();", adv_search_btn)
        time.sleep(2)
        
        # Asegurarse de que estamos en la pestaña de Batch Search
        print("Seleccionando pestaña de Batch Search...")
        batch_tab = wait.until(
            EC.presence_of_element_located((By.XPATH, "//ul[@id='tabSearchType']//li[@data-id='2']//a"))
        )
        if not "active" in batch_tab.find_element(By.XPATH, "..").get_attribute("class"):
            driver.execute_script("arguments[0].click();", batch_tab)
            time.sleep(1)
        
        # Preparar los valores de CODIGO_LOTE con el formato requerido
        search_text = "\n".join(search_codes)
        print(f"Preparando {len(search_codes)} códigos para búsqueda...")
        
        # Encontrar y llenar el textarea dentro del panel activo
        print("Ingresando códigos en el formulario...")
        textarea = wait.until(
            EC.presence_of_element_located((By.XPATH, "//div[@id='tab_15_2']//textarea[@id='adsBatch']"))
        )
        driver.execute_script("arguments[0].value = arguments[1];", textarea, search_text)
        time.sleep(1)
        
        # Hacer clic en el botón de búsqueda
        print("Iniciando búsqueda...")
        search_button = wait.until(
            EC.element_to_be_clickable((By.ID, "btnAdvancedSearch"))
        )
        driver.execute_script("arguments[0].click();", search_button)
        print("Búsqueda iniciada")
        
        # Esperar según la cantidad de registros
        wait_for_results(driver, len(search_codes))
        
        # Seleccionar y descargar todos los archivos
        select_and_download_all_files(driver)
        
    except Exception as e:
        print(f"Error durante la búsqueda avanzada: {str(e)}")
        print("Detalles del error:")
        try:
            print(f"URL actual: {driver.current_url}")
            print("Estado de los elementos:")
            print(f"- Modal visible: {driver.find_element(By.CLASS_NAME, 'modal-dialog').is_displayed()}")
            print(f"- Pestaña Batch Search activa: {driver.find_element(By.XPATH, "//li[@data-id='2']").get_attribute('class')}")
            print(f"- Panel Batch Search activo: {driver.find_element(By.ID, 'tab_15_2').get_attribute('class')}")
            print(f"- Textarea visible: {driver.find_element(By.XPATH, "//div[@id='tab_15_2']//textarea[@id='adsBatch']").is_displayed()}")
            print(f"- Botón buscar visible: {driver.find_element(By.ID, 'btnAdvancedSearch').is_displayed()}")
        except Exception as detail_error:
            print(f"Error al obtener detalles: {str(detail_error)}")

def extract_pdf_info(pdf_path):
    """Extraer información del PDF: Material Code, Lot No y Expiry Date"""
    try:
        reader = PdfReader(pdf_path)
        text = reader.pages[0].extract_text()
        
        # Patrones para buscar la información
        material_code_pattern = r"Material Code\s*(\d{3}-\d{6}-\d{2})"
        lot_no_pattern = r"Lot No\.\s*(\d+)"
        expiry_date_pattern = r"Expiry date\s*(\d{4}-\d{2}-\d{2})"
        
        # Extraer información usando regex
        material_code = re.search(material_code_pattern, text)
        lot_no = re.search(lot_no_pattern, text)
        expiry_date = re.search(expiry_date_pattern, text)
        
        if material_code and lot_no and expiry_date:
            return {
                'material_code': material_code.group(1),
                'lot_no': lot_no.group(1),
                'expiry_date': expiry_date.group(1)
            }
        else:
            print(f"No se pudo extraer toda la información del PDF: {pdf_path}")
            return None
            
    except Exception as e:
        print(f"Error al procesar el PDF {pdf_path}: {str(e)}")
        return None

def update_excel_with_pdf_data():
    """Actualizar el Excel original con la información de los PDFs"""
    try:
        print("\nActualizando Excel con información de los PDFs...")
        
        # Rutas de archivos
        excel_path = BASE_DIR / EXCEL_FILE
        backup_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f'SEGUIMIENTO_ALMACEN_actualizado_{backup_timestamp}.xlsx'
        backup_path = BASE_DIR / backup_file
        
        # Leer el Excel original preservando el formato
        print("Leyendo archivo Excel original...")
        for attempt in range(3):  # Intentar 3 veces
            try:
                # Primero hacer una copia del archivo original para preservar formato
                shutil.copy2(excel_path, backup_path)
                
                # Primero obtener el nombre de la hoja
                wb_temp = pd.ExcelFile(excel_path)
                sheet_name = wb_temp.sheet_names[0]  # Usar la primera hoja
                print(f"Usando hoja: {sheet_name}")
                
                # Leer el Excel con el nombre de hoja correcto
                df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
                
                # Verificar las columnas disponibles
                print("\nColumnas en el Excel:")
                for idx, col in enumerate(df.columns):
                    print(f"Índice {idx}: {col}")
                
                break
            except PermissionError:
                if attempt < 2:
                    print("Excel está abierto. Esperando 5 segundos...")
                    time.sleep(5)
                else:
                    raise
        
        # Obtener índices de columnas (base 0)
        codigo_col = excel_column_to_number(EXCEL_COL_CODIGO) - 1
        lote_col = excel_column_to_number(EXCEL_COL_LOTE) - 1
        fecha_venc_col = excel_column_to_number(EXCEL_COL_FECHA_VENC) - 1
        
        print(f"Usando columna {EXCEL_COL_CODIGO} (índice {codigo_col}) para código")
        print(f"Usando columna {EXCEL_COL_LOTE} (índice {lote_col}) para lote")
        print(f"Usando columna {EXCEL_COL_FECHA_VENC} (índice {fecha_venc_col}) para fecha de vencimiento")
        
        # Contador de actualizaciones
        updates = 0
        updated_rows = []  # Lista para guardar las filas actualizadas
        
        # Procesar cada PDF en el directorio de descargas
        for pdf_file in DOWNLOADS_DIR.glob("*.pdf"):
            pdf_info = extract_pdf_info(pdf_file)
            
            if pdf_info:
                # Crear el código combinado (material_code + lot_no)
                combined_code = f"{pdf_info['material_code']} {pdf_info['lot_no']}"
                
                # Crear el código combinado en el DataFrame
                df_combined = df.iloc[:, codigo_col].astype(str).str.replace('MR ', '') + ' ' + df.iloc[:, lote_col].astype(str)
                
                # Buscar coincidencias usando el código combinado
                matches = df_combined == combined_code
                
                if matches.any():
                    # Actualizar la columna de fecha de vencimiento con la fecha del PDF
                    df.iloc[matches, fecha_venc_col] = pdf_info['expiry_date']
                    updates += 1
                    updated_rows.extend(df.index[matches].tolist())
                    print(f"Actualizado: {combined_code} con fecha {pdf_info['expiry_date']}")
        
        if updates > 0:
            print("\nGuardando cambios...")
            try:
                # Leer el archivo de backup (que tiene el formato original)
                from openpyxl import load_workbook
                wb = load_workbook(backup_path)
                ws = wb[sheet_name]  # Usar el nombre de hoja correcto
                
                print(f"\nActualizando celdas en columna {EXCEL_COL_FECHA_VENC}...")
                # Actualizar solo las celdas específicas con las nuevas fechas
                for row_idx in updated_rows:
                    cell_ref = f"{EXCEL_COL_FECHA_VENC}{row_idx + 2}"  # +2 porque Excel es 1-based y hay encabezado
                    print(f"Actualizando celda {cell_ref}")
                    ws[cell_ref] = df.iloc[row_idx, fecha_venc_col]
                
                # Guardar el backup con los cambios
                wb.save(backup_path)
                
                print(f"\nProceso completado:")
                print(f"- {updates} registros actualizados")
                print(f"- Archivo actualizado guardado como: {backup_file}")
                print(f"- El archivo original {EXCEL_FILE} no ha sido modificado")
                
            except Exception as e:
                print(f"Error al guardar los cambios: {str(e)}")
                print("Detalles del error:")
                import traceback
                traceback.print_exc()
                raise
        else:
            print("\nNo se encontraron coincidencias para actualizar")
            # Eliminar el archivo de backup si no hubo cambios
            if backup_path.exists():
                backup_path.unlink()
        
    except Exception as e:
        print(f"Error al actualizar el Excel: {str(e)}")
        print("\nSugerencias:")
        print("1. Asegúrate de que el archivo Excel no esté abierto en otro programa")
        print("2. Verifica que tienes permisos de escritura en la carpeta")
        print(f"3. Verifica que la hoja '{sheet_name}' existe en el archivo")
        print(f"4. Verifica que las columnas {EXCEL_COL_CODIGO}, {EXCEL_COL_LOTE} y {EXCEL_COL_FECHA_VENC} existen y son accesibles")
        raise

def main():
    try:
        # Verificar credenciales
        if not MINDRAY_USERNAME or not MINDRAY_PASSWORD:
            raise ValueError("Error: Las credenciales de Mindray no están configuradas en el archivo .env")
        
        print(f"\nUsando directorio de descargas: {PDF_DIR_NAME}")
        
        # Asegurar que el directorio existe
        DOWNLOADS_DIR.mkdir(exist_ok=True)
        print(f"Ruta completa: {DOWNLOADS_DIR}")
        
        # Obtener lista de códigos a buscar
        search_codes = read_excel_data()
        print(f"\nSe buscarán {len(search_codes)} códigos")
        
        # Iniciar proceso de automatización web
        driver = setup_driver()
        try:
            login_successful = login_mindray(driver)
            
            if login_successful:
                print("\nIniciando proceso de búsqueda...")
                perform_advanced_search(driver, search_codes)
                print("\nDescarga de archivos completada. Cerrando navegador...")
        finally:
            driver.quit()
            
        # Procesar archivos ZIP después de cerrar el navegador
        print("\nProcesando todos los archivos ZIP descargados...")
        process_zip_files()
        
        # Actualizar Excel con la información de los PDFs
        print("\nProcesando PDFs y actualizando Excel...")
        update_excel_with_pdf_data()
        
        print("\nProceso completado exitosamente!")
            
    except Exception as e:
        print(f"\n=== ERROR ===")
        print(str(e))
        raise

if __name__ == "__main__":
    main() 